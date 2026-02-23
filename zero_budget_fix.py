import openpyxl
import os

source_path = r'c:\Users\past9\OneDrive\Desktop\project\DASHBOARD\Strat_Edge_Import_Ready.xlsx'
target_path = r'c:\Users\past9\OneDrive\Desktop\project\DASHBOARD\Strat_Edge_Zero_Budget.xlsx'

wb = openpyxl.load_workbook(source_path)
ws1 = wb['Project_Schedule']

# 1. Set Total Budget to 0
ws1['F5'] = 0.0
# Column F is index 6
for r in range(12, 20):
    val = ws1.cell(r, 6).value
    if val is not None:
        ws1.cell(r, 6).value = 0.0

if 'Expenditure_Log' in wb.sheetnames:
    ws2 = wb['Expenditure_Log']
    if ws2.max_row >= 5:
        ws2.delete_rows(5, ws2.max_row - 4)

wb.save(target_path)
print(f"Created: {target_path}")
