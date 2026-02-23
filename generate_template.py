"""
Professional Excel Template Generator
Creates a polished, company-grade template matching the user's existing format.
Updated: Added 'Responsible (Username)' and 'Expected Output' columns.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime, timedelta
import os

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'templates')
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Styling Constants
HEADER_FILL = PatternFill(start_color="2C5AA0", end_color="2C5AA0", fill_type="solid")  # Dark Blue
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
TITLE_FONT = Font(bold=True, size=14)
LABEL_FONT = Font(bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_header_row(ws, row_num, start_col, end_col):
    """Apply header styling to a row."""
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

def create_professional_template(sample_data=False):
    """Creates a professional Excel template."""
    
    wb = Workbook()
    
    # =========================================================================
    # SHEET 1: PROJECT INFO & SCHEDULE (Gantt as Table)
    # =========================================================================
    ws1 = wb.active
    ws1.title = "Project_Schedule"
    
    # --- HEADER SECTION ---
    ws1.merge_cells('A1:J1')
    ws1['A1'] = "PROJECT STATUS DASHBOARD - DATA ENTRY FORM"
    ws1['A1'].font = Font(bold=True, size=16, color="2C5AA0")
    ws1['A1'].alignment = Alignment(horizontal='center')
    
    # Period
    ws1['A3'] = "REPORTING PERIOD:"
    ws1['A3'].font = LABEL_FONT
    ws1['B3'] = datetime.now().strftime("%Y/%m") if sample_data else ""
    ws1['B3'].fill = PatternFill(start_color="E6F0FA", end_color="E6F0FA", fill_type="solid")
    
    # Project Info Block (Column B for values)
    labels = [
        ('A5', 'PROJECT NAME:'),
        ('A6', 'PROJECT NUMBER:'),
        ('A7', 'CLIENT:'),
        ('A8', 'PROJECT MANAGER:'),
    ]
    values_sample = [
        ('B5', 'Liberty Tower Fibre Relocation'),
        ('B6', 'IN963051'),
        ('B7', 'Liberty Properties'),
        ('B8', 'J. Smith'),
    ]
    
    for (cell, label) in labels:
        ws1[cell] = label
        ws1[cell].font = LABEL_FONT
        ws1[cell].border = THIN_BORDER
    
    if sample_data:
        for (cell, val) in values_sample:
            ws1[cell] = val
            ws1[cell].border = THIN_BORDER
    
    # Financial Summary Block (Column F for values)
    ws1['E5'] = "TOTAL BUDGET:"
    ws1['E6'] = "START DATE:"
    ws1['E7'] = "TARGET END DATE:"
    ws1['E8'] = "STATUS:"
    for row in range(5, 9):
        ws1.cell(row=row, column=5).font = LABEL_FONT
        ws1.cell(row=row, column=5).border = THIN_BORDER
        ws1.cell(row=row, column=6).border = THIN_BORDER
    
    if sample_data:
        ws1['F5'] = 240100.82
        ws1['F5'].number_format = 'R #,##0.00'
        ws1['F6'] = datetime(2025, 1, 15)
        ws1['F6'].number_format = 'YYYY-MM-DD'
        ws1['F7'] = datetime(2025, 6, 30)
        ws1['F7'].number_format = 'YYYY-MM-DD'
        ws1['F8'] = "In Progress"
        ws1['F8'].fill = SECTION_FILL
    
    # --- PROJECT SCHEDULE TABLE (MATCHES REFERENCE FILE) ---
    ws1['A10'] = "PROJECT SCHEDULE (BASELINE PLAN)"
    ws1['A10'].font = TITLE_FONT
    
    # Precise Headers from Strat Edge-Project Management Tool.xlsx
    schedule_headers = [
        'Activity ID', 'Activity ', 'Output', 'Planned Start', 'Planned End',
        'Budgeted Cost (R)', 'Depends On', 'Actual Start', 'Actual End',
        'Person Responsible'
    ]
    for col, header in enumerate(schedule_headers, start=1):
        ws1.cell(row=11, column=col, value=header)
    style_header_row(ws1, 11, 1, len(schedule_headers))
    
    # Sample data rows (Reordered to match sequence above)
    if sample_data:
        schedule_data = [
            # ID, Name, Output, Start, End, Budget, Depends, ActStart, ActEnd, Resp
            [1, 'Site Survey & Mobilization', 'Survey Report PDF', datetime(2025,1,15), datetime(2025,1,20), 15000, '-', datetime(2025,1,15), datetime(2025,1,19), 'pm_user'],
            [2, 'Civil Works - Trenching', 'Trench Completion Certificate', datetime(2025,1,21), datetime(2025,2,15), 45000, 1, datetime(2025,1,21), datetime(2025,2,18), 'recorder'],
            [3, 'Fibre Cable Laying', 'Cable Installation Report', datetime(2025,2,1), datetime(2025,3,15), 60000, 2, datetime(2025,2,5), None, 'pm_user'],
            [4, 'Fibre Splicing', 'Splice Test Results', datetime(2025,3,16), datetime(2025,4,15), 35000, 3, None, None, 'recorder'],
            [5, 'Testing & Commissioning', 'Commissioning Sign-Off', datetime(2025,4,16), datetime(2025,5,15), 25000, 4, None, None, 'pm_user'],
            [6, 'Site Restoration & Handover', 'Handover Certificate', datetime(2025,5,16), datetime(2025,6,15), 20000, 5, None, None, 'pm_user'],
        ]
        for row_idx, row_data in enumerate(schedule_data, start=12):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws1.cell(row=row_idx, column=col_idx, value=value)
                cell.border = THIN_BORDER
                if col_idx == 6 and value:
                    cell.number_format = 'R #,##0.00'
                if isinstance(value, datetime):
                    cell.number_format = 'YYYY-MM-DD'
    else:
        # Empty rows for data entry
        for row_idx in range(12, 22):
            for col_idx in range(1, len(schedule_headers) + 1):
                ws1.cell(row=row_idx, column=col_idx).border = THIN_BORDER
    
    # Adjust column widths
    ws1.column_dimensions['A'].width = 12
    ws1.column_dimensions['B'].width = 30 # Activity
    ws1.column_dimensions['C'].width = 30 # Output
    ws1.column_dimensions['D'].width = 14 # Planned Start
    ws1.column_dimensions['E'].width = 14 # Planned End
    ws1.column_dimensions['F'].width = 18 # Budgeted Cost
    ws1.column_dimensions['G'].width = 12 # Depends On
    ws1.column_dimensions['H'].width = 14 # Actual Start
    ws1.column_dimensions['I'].width = 14 # Actual End
    ws1.column_dimensions['J'].width = 22 # Person Responsible
    
    # =========================================================================
    # SHEET 2: EXPENDITURE LOG
    # =========================================================================
    ws2 = wb.create_sheet(title="Expenditure_Log")
    
    ws2['A1'] = "EXPENDITURE LOG"
    ws2['A1'].font = TITLE_FONT
    ws2['A2'] = "Record every payment made. Attach invoice/PO reference."
    ws2['A2'].font = Font(italic=True, color="666666")
    
    exp_headers = ['Date', 'Activity ID', 'Category', 'Description', 'Reference (Invoice/PO)', 'Amount (R)']
    for col, header in enumerate(exp_headers, start=1):
        ws2.cell(row=4, column=col, value=header)
    style_header_row(ws2, 4, 1, len(exp_headers))
    
    if sample_data:
        exp_data = [
            [datetime(2025,1,17), 1, 'Labour', 'Survey team wages', 'INV-001', 8500.00],
            [datetime(2025,1,18), 1, 'Vehicle', 'Site transport', 'VEH-010', 2500.00],
            [datetime(2025,1,25), 2, 'Labour', 'Trenching crew - Week 1', 'INV-005', 12000.00],
            [datetime(2025,2,1), 2, 'Labour', 'Trenching crew - Week 2', 'INV-008', 11500.00],
            [datetime(2025,2,5), 2, 'Material', 'Ducting pipes', 'PO-101', 18000.00],
            [datetime(2025,2,10), 3, 'Material', 'Fibre cable - 500m', 'PO-115', 25000.00],
            [datetime(2025,2,15), 2, 'Vehicle', 'Excavator rental', 'VEH-022', 4500.00],
            [datetime(2025,2,20), 3, 'Labour', 'Cable laying team', 'INV-012', 9500.00],
        ]
        for row_idx, row_data in enumerate(exp_data, start=5):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws2.cell(row=row_idx, column=col_idx, value=value)
                cell.border = THIN_BORDER
                if col_idx == 1:
                    cell.number_format = 'YYYY-MM-DD'
                if col_idx == 6:
                    cell.number_format = 'R #,##0.00'
    else:
        for row_idx in range(5, 30):
            for col_idx in range(1, len(exp_headers) + 1):
                ws2.cell(row=row_idx, column=col_idx).border = THIN_BORDER
    
    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 12
    ws2.column_dimensions['D'].width = 30
    ws2.column_dimensions['E'].width = 22
    ws2.column_dimensions['F'].width = 15
    
    # =========================================================================
    # SHEET 3: RISKS & ISSUES
    # =========================================================================
    ws3 = wb.create_sheet(title="Risk_Register")
    
    ws3['A1'] = "RISK & ISSUES REGISTER"
    ws3['A1'].font = TITLE_FONT
    
    risk_headers = ['Date Identified', 'Risk/Issue Description', 'Impact (H/M/L)', 'Status', 'Mitigation Action']
    for col, header in enumerate(risk_headers, start=1):
        ws3.cell(row=3, column=col, value=header)
    style_header_row(ws3, 3, 1, len(risk_headers))
    
    if sample_data:
        risk_data = [
            [datetime(2025,1,22), 'Trenching delayed due to unexpected rock', 'H', 'Resolved', 'Rented rock breaker'],
            [datetime(2025,2,10), 'Material delivery delayed by supplier', 'M', 'Open', 'Escalated to procurement'],
        ]
        for row_idx, row_data in enumerate(risk_data, start=4):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws3.cell(row=row_idx, column=col_idx, value=value)
                cell.border = THIN_BORDER
                if col_idx == 1:
                    cell.number_format = 'YYYY-MM-DD'
    else:
        for row_idx in range(4, 15):
            for col_idx in range(1, len(risk_headers) + 1):
                ws3.cell(row=row_idx, column=col_idx).border = THIN_BORDER
    
    ws3.column_dimensions['A'].width = 15
    ws3.column_dimensions['B'].width = 40
    ws3.column_dimensions['C'].width = 15
    ws3.column_dimensions['D'].width = 12
    ws3.column_dimensions['E'].width = 35
    
    # =========================================================================
    # SAVE
    # =========================================================================
    suffix = "_Sample_v2" if sample_data else ""
    filename = f"Project_Template{suffix}.xlsx"
    output_path = os.path.join(OUTPUT_DIR, filename)
    wb.save(output_path)
    print(f"✅ Created: {output_path}")
    return output_path

# =============================================================================
# MAIN
# =============================================================================
if __name__ == '__main__':
    print("=" * 60)
    print("GENERATING PROFESSIONAL PROJECT TEMPLATES")
    print("=" * 60)
    
    create_professional_template(sample_data=False)  # Blank form
    create_professional_template(sample_data=True)   # With sample data
    
    print("\n" + "=" * 60)
    print("DONE! Two files created:")
    print("  1. Project_Template.xlsx        (Blank form for new projects)")
    print("  2. Project_Template_Sample_v2.xlsx (Example with Liberty Towers data)")
    print("=" * 60)
