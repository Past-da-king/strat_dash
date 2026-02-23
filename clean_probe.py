import openpyxl

source_file = r'c:\Users\past9\OneDrive\Desktop\project\DASHBOARD\Strat Edge-Project Management Tool copy.xlsx'
wb = openpyxl.load_workbook(source_file, data_only=True)
ws = wb['Project_Schedule']

for r in range(5, 10):
    for c in range(1, 10):
        v = ws.cell(r, c).value
        if v:
            print(f"Cell {openpyxl.utils.get_column_letter(c)}{r}: '{v}'")
