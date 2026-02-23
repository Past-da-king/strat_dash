import openpyxl

file_path = 'Strat Edge-Project Management Tool.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)

ws = wb['Project_Schedule']
print("--- Project Info ---")
for r in range(5, 9):
    print(f"Row {r}: A:{ws.cell(r, 1).value} | B:{ws.cell(r, 2).value} | E:{ws.cell(r, 5).value} | F:{ws.cell(r, 6).value}")

print("\n--- Project Schedule Data ---")
for r in range(11, 20):
    row = [ws.cell(r, c).value for c in range(1, 11)]
    if any(row):
        print(f"Row {r}: {row}")

ws_exp = wb['Expenditure_Log']
print("\n--- Expenditure Log Data ---")
for r in range(4, 10):
    row = [ws_exp.cell(r, c).value for c in range(1, 7)]
    if any(row):
        print(f"Row {r}: {row}")

ws_risk = wb['Risk_Register']
print("\n--- Risk Register Data ---")
for r in range(3, 8):
    row = [ws_risk.cell(r, c).value for c in range(1, 6)]
    if any(row):
        print(f"Row {r}: {row}")
