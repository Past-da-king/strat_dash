import openpyxl

file_path = 'Strat Edge-Project Management Tool.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)

for sheet_name in wb.sheetnames:
    print(f"\n===== Sheet: {sheet_name} =====")
    ws = wb[sheet_name]
    for row in ws.iter_rows(values_only=True):
        if any(row):
            print(row)
