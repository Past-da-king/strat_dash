import openpyxl
from openpyxl.utils import get_column_letter

file_path = 'Strat Edge-Project Management Tool.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)

def scan_sheet(sheet_name, max_row=30, max_col=12):
    print(f"\n===== SCANNING SHEET: {sheet_name} =====")
    ws = wb[sheet_name]
    for r in range(1, max_row + 1):
        row_vals = []
        for c in range(1, max_col + 1):
            val = ws.cell(row=r, column=c).value
            if val is not None:
                row_vals.append(f"{get_column_letter(c)}{r}: {val}")
        if row_vals:
            print(" | ".join(row_vals))

scan_sheet('Project_Schedule')
scan_sheet('Expenditure_Log')
scan_sheet('Risk_Register')
