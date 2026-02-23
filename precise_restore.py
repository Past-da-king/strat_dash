import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime

source_file = r'c:\Users\past9\OneDrive\Desktop\project\DASHBOARD\Strat Edge-Project Management Tool copy.xlsx'
output_file = r'c:\Users\past9\OneDrive\Desktop\project\DASHBOARD\Strat_Edge_PRECISE_v2.xlsx'

# Load the user's real project file (source of truth)
wb_src = openpyxl.load_workbook(source_file, data_only=True)
wb_out = openpyxl.Workbook()

# Style constants
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

def copy_sheet_precisely(src_ws_name, out_ws):
    src_ws = wb_src[src_ws_name]
    for r in range(1, src_ws.max_row + 1):
        for c in range(1, src_ws.max_column + 1):
            source_cell = src_ws.cell(r, c)
            target_cell = out_ws.cell(r, c, source_cell.value)
            target_cell.border = THIN_BORDER

ws_out = wb_out.active
ws_out.title = 'Project_Schedule'

# 1. Copy everything from original
copy_sheet_precisely('Project_Schedule', ws_out)

# 2. CLEAR the problematic area to ensure no duplicates
# We want Labels in A/E and Values in B/F. 
# The original had labels in A and F. We need to move F labels to E.
for r in range(5, 9):
    # Clear E and F to be sure
    ws_out.cell(r, 5).value = None 
    ws_out.cell(r, 6).value = None

# 3. SET LABELS PROPERLY (Column A and Column E)
ws_out['A5'] = "PROJECT NAME:"
ws_out['A6'] = "PROJECT NUMBER:"
ws_out['A7'] = "CLIENT:"
ws_out['A8'] = "PROJECT MANAGER:"

ws_out['E5'] = "TOTAL BUDGET:"
ws_out['E6'] = "START DATE:"
ws_out['E7'] = "TARGET END DATE:"
ws_out['E8'] = "STATUS:"

# 4. SET VALUES PROPERLY (Project values from B, and Budget/Dates into F)
# (Project values are already in B from the copy, but let's re-verify)

# Set Budgets/Dates into F (which we just cleared)
ws_out['F5'] = 0.0
ws_out['F5'].number_format = '\"R\" #,##0.00'
ws_out['F6'] = "2026-03-01" # Start date from user data
ws_out['F7'] = "2026-08-30" # End date from user data
ws_out['F8'] = "Active"

# 5. ZERO OUT budgets for all activities (Column 6)
for r in range(12, ws_out.max_row + 1):
    if ws_out.cell(r, 2).value: # If there is an activity name
        ws_out.cell(r, 6).value = 0.0
        ws_out.cell(r, 6).number_format = '\"R\" #,##0.00'

# 6. Copy other sheets
if 'Expenditure_Log' in wb_src.sheetnames:
    ws_exp = wb_out.create_sheet('Expenditure_Log')
    copy_sheet_precisely('Expenditure_Log', ws_exp)
    # Zero out amounts
    for r in range(5, ws_exp.max_row + 1):
        if ws_exp.cell(r, 6).value:
            ws_exp.cell(r, 6).value = 0.0

if 'Risk_Register' in wb_src.sheetnames:
    ws_risk = wb_out.create_sheet('Risk_Register')
    copy_sheet_precisely('Risk_Register', ws_risk)

# formatting
ws_out.column_dimensions['B'].width = 45
ws_out.column_dimensions['C'].width = 35
ws_out.column_dimensions['F'].width = 20

wb_out.save(output_file)
print(f"Successfully restored and fixed metadata alignment in: {output_file}")
