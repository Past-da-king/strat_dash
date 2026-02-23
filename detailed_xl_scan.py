import pandas as pd
import openpyxl

file_path = 'Strat Edge-Project Management Tool.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)

for sheet_name in wb.sheetnames:
    print(f"\n===== Sheet: {sheet_name} =====")
    ws = wb[sheet_name]
    for row in range(1, 16):
        row_vals = []
        for col in range(1, 15):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                row_vals.append(f"[{openpyxl.utils.get_column_letter(col)}{row}]: {val}")
        if row_vals:
            print(" | ".join(row_vals))
