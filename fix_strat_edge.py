import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime, date

file_path = 'Strat Edge-Project Management Tool.xlsx'
wb = openpyxl.load_workbook(file_path)

# Styling
HEADER_FILL = PatternFill(start_color="2C5AA0", end_color="2C5AA0", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

def style_range(ws, row, start_col, end_col):
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

# 1. FIX PROJECT_SCHEDULE
ws1 = wb['Project_Schedule']

# Metadata (Col B and Col F)
ws1['B5'] = "STRAT EDGE LINKAGES & PARTNERSHIPS"
ws1['B6'] = "SE-PRJ-2026-01"
ws1['B7'] = "Strat Edge Group"
ws1['B8'] = "Admin"

ws1['F5'] = 750000.00
ws1['F5'].number_format = 'R #,##0.00'
ws1['F6'] = datetime(2026, 3, 1)
ws1['F6'].number_format = 'YYYY-MM-DD'
ws1['F7'] = datetime(2026, 8, 30)
ws1['F7'].number_format = 'YYYY-MM-DD'
ws1['F8'] = "Active"

# Schedule Table Headers (Activity ID, Activity, Output, Start, End, Budget, Depends, ActStart, ActEnd, Resp)
headers = ['Activity ID', 'Activity ', 'Output', 'Planned Start', 'Planned End', 'Budgeted Cost (R)', 'Depends On', 'Actual Start', 'Actual End', 'Person Responsible']
for c, h in enumerate(headers, 1):
    ws1.cell(row=11, column=c, value=h)
style_range(ws1, 11, 1, 10)

# Fill Activities
activities = [
    [1, 'Stakeholder Analysis & Mapping', 'Stakeholder Matrix', datetime(2026,3,1), datetime(2026,3,10), 25000, '-', datetime(2026,3,2), None, 'admin'],
    [2, 'Partnership Framework Design', 'Framework Document', datetime(2026,3,11), datetime(2026,4,5), 85000, 1, None, None, 'pm_user'],
    [3, 'Linkage Strategy Development', 'Strategy Report', datetime(2026,4,6), datetime(2026,5,15), 120000, 2, None, None, 'pm_user'],
    [4, 'Compliance & Internal Submission', 'Compliance Certificate', datetime(2026,5,16), datetime(2026,6,10), 45000, 3, None, None, 'recorder'],
    [5, 'Implementation of Shared Platform', 'Platform Go-Live', datetime(2026,6,11), datetime(2026,7,30), 250000, 4, None, None, 'pm_user'],
    [6, 'Final Review & Handover', 'Contract Completion', datetime(2026,8,1), datetime(2026,8,30), 50000, 5, None, None, 'admin']
]

# Clear existing rows first (from row 12)
for r in range(12, 30):
    for c in range(1, 11):
        ws1.cell(r, c).value = None
        ws1.cell(r, c).border = None

for i, act in enumerate(activities):
    r = 12 + i
    for c, val in enumerate(act, 1):
        cell = ws1.cell(r, c, val)
        cell.border = THIN_BORDER
        if isinstance(val, datetime):
            cell.number_format = 'YYYY-MM-DD'
        if c == 6:
            cell.number_format = 'R #,##0.00'

# 2. FIX EXPENDITURE_LOG
if 'Expenditure_Log' not in wb.sheetnames: wb.create_sheet('Expenditure_Log')
ws2 = wb['Expenditure_Log']
ws2.delete_rows(1, 100)
ws2['A1'] = "EXPENDITURE LOG"
ws2['A1'].font = Font(bold=True, size=14)
exp_headers = ['Date', 'Activity ID', 'Category', 'Description', 'Reference (Invoice/PO)', 'Amount (R)']
for c, h in enumerate(exp_headers, 1):
    ws2.cell(row=4, column=c, value=h)
style_range(ws2, 4, 1, 6)

expenses = [
    [datetime(2026,3,5), 1, 'Consultancy', 'Stakeholder meeting fees', 'INV-881', 12000.00],
    [datetime(2026,3,8), 1, 'Travel', 'Regional site visits', 'EXP-042', 4500.00],
]
for i, exp in enumerate(expenses):
    r = 5 + i
    for c, val in enumerate(exp, 1):
        cell = ws2.cell(r, c, val)
        cell.border = THIN_BORDER
        if c == 1: cell.number_format = 'YYYY-MM-DD'
        if c == 6: cell.number_format = 'R #,##0.00'

# 3. FIX RISK_REGISTER
if 'Risk_Register' not in wb.sheetnames: wb.create_sheet('Risk_Register')
ws3 = wb['Risk_Register']
ws3.delete_rows(1, 100)
ws3['A1'] = "RISK & ISSUES REGISTER"
ws3['A1'].font = Font(bold=True, size=14)
risk_headers = ['Date Identified', 'Risk/Issue Description', 'Impact (H/M/L)', 'Status', 'Mitigation Action']
for c, h in enumerate(risk_headers, 1):
    ws3.cell(row=3, column=c, value=h)
style_range(ws3, 3, 1, 5)

risks = [
    [datetime(2026,3,1), 'Delay in stakeholder responses', 'M', 'Open', 'Send follow-up reminders'],
    [datetime(2026,3,4), 'Compliance framework update', 'H', 'Open', 'Consult with legal team'],
]
for i, risk in enumerate(risks):
    r = 4 + i
    for c, val in enumerate(risk, 1):
        cell = ws3.cell(r, c, val)
        cell.border = THIN_BORDER
        if c == 1: cell.number_format = 'YYYY-MM-DD'

# Final formatting widths
ws1.column_dimensions['A'].width = 12
ws1.column_dimensions['B'].width = 35
ws1.column_dimensions['C'].width = 30
ws1.column_dimensions['D'].width = 15
ws1.column_dimensions['E'].width = 15
ws1.column_dimensions['F'].width = 18
ws1.column_dimensions['G'].width = 12
ws1.column_dimensions['H'].width = 15
ws1.column_dimensions['I'].width = 15
ws1.column_dimensions['J'].width = 22

wb.save(file_path)
print(f"Successfully fixed and populated {file_path}")
