import openpyxl

source_file = r'c:\Users\past9\OneDrive\Desktop\project\DASHBOARD\Strat Edge-Project Management Tool copy.xlsx'
wb = openpyxl.load_workbook(source_file, data_only=True)
ws = wb['Project_Schedule']

for r in range(5, 10):
    row_data = []
    for c in range(1, 15):
        v = ws.cell(r, c).value
        row_data.append(f"{openpyxl.utils.get_column_letter(c)}: {v}")
    print(f"Row {r} | " + " | ".join(row_data))
