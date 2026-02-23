import openpyxl
import json

file_path = 'Strat Edge-Project Management Tool.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)

data = {}

for sn in wb.sheetnames:
    ws = wb[sn]
    sheet_data = []
    for row in ws.iter_rows(values_only=True):
        if any(v is not None for v in row):
            sheet_data.append([str(v) if v is not None else None for v in row])
    data[sn] = sheet_data

print(json.dumps(data, indent=2))
