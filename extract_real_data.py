import openpyxl

source_file = r'c:\Users\past9\OneDrive\Desktop\project\DASHBOARD\Strat Edge-Project Management Tool copy.xlsx'
wb = openpyxl.load_workbook(source_file, data_only=True)
ws = wb['Project_Schedule']

print("--- START OF SCHEDULE ---")
for r in range(12, 40): # Capturing up to 40 rows to be safe
    row = [ws.cell(r, c).value for c in range(1, 15)]
    if any(row):
        print(f"Row {r}: {row}")
print("--- END OF SCHEDULE ---")

print("\n--- METADATA ---")
for r in range(5, 9):
    print(f"Row {r} Col B: {ws.cell(r, 2).value} | Col F: {ws.cell(r, 6).value}")
