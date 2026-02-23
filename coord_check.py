import openpyxl

source_file = r'c:\Users\past9\OneDrive\Desktop\project\DASHBOARD\Strat Edge-Project Management Tool copy.xlsx'
wb = openpyxl.load_workbook(source_file, data_only=True)
ws = wb['Project_Schedule']

print("--- METADATA COORDS ---")
for r in range(5, 10):
    row_str = []
    for c in range(1, 10):
        val = ws.cell(r, c).value
        if val:
            row_str.append(f"[{openpyxl.utils.get_column_letter(c)}{r}]: '{val}'")
    print(" | ".join(row_str))
