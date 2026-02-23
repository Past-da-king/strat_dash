import openpyxl
import json

source_file = r'c:\Users\past9\OneDrive\Desktop\project\DASHBOARD\Strat Edge-Project Management Tool copy.xlsx'
wb = openpyxl.load_workbook(source_file, data_only=True)

data = {}
for sn in wb.sheetnames:
    ws = wb[sn]
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(row)
    data[sn] = rows

print(json.dumps(data, default=str))
