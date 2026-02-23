import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime

# We will create a NEW file to avoid overwriting the original
original_file = 'Strat Edge-Project Management Tool.xlsx'
output_file = 'Strat_Edge_Import_Ready.xlsx'

# Try to load the original to keep any extra sheets, or start fresh if needed
try:
    wb = openpyxl.load_workbook(original_file)
except:
    wb = openpyxl.Workbook()

if 'Project_Schedule' not in wb.sheetnames:
    ws1 = wb.active
    ws1.title = 'Project_Schedule'
else:
    ws1 = wb['Project_Schedule']

# Styling
HEADER_FILL = PatternFill(start_color="2C5AA0", end_color="2C5AA0", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
LABEL_FONT = Font(bold=True, size=11)
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

def style_header(ws, r, c_start, c_end):
    for c in range(c_start, c_end + 1):
        cell = ws.cell(r, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

# --- 1. PROJECT INFO (LABELS in A/E, VALUES in B/F) ---
# Column A Labels
ws1['A5'] = "PROJECT NAME:"
ws1['A6'] = "PROJECT NUMBER:"
ws1['A7'] = "CLIENT:"
ws1['A8'] = "PROJECT MANAGER:"

# Column B Values
ws1['B5'] = "STRAT EDGE LINKAGES & PARTNERSHIPS"
ws1['B6'] = "SE-PRJ-2026-01"
ws1['B7'] = "Strat Edge Group"
ws1['B8'] = "Admin"

# Column E Labels
ws1['E5'] = "TOTAL BUDGET:"
ws1['E6'] = "START DATE:"
ws1['E7'] = "TARGET END DATE:"
ws1['E8'] = "STATUS:"

# Column F Values
ws1['F5'] = 750000.00
ws1['F5'].number_format = 'R #,##0.00'
ws1['F6'] = datetime(2026, 3, 1)
ws1['F6'].number_format = 'YYYY-MM-DD'
ws1['F7'] = datetime(2026, 8, 30)
ws1['F7'].number_format = 'YYYY-MM-DD'
ws1['F8'] = "Active"

# Apply styling to info block
for row in range(5, 9):
    for col in [1, 2, 5, 6]:
        cell = ws1.cell(row, col)
        cell.border = THIN_BORDER
        if col in [1, 5]:
            cell.font = LABEL_FONT

# --- 2. SCHEDULE TABLE ---
schedule_headers = [
    'Activity ID', 'Activity ', 'Output', 'Planned Start', 'Planned End',
    'Budgeted Cost (R)', 'Depends On', 'Actual Start', 'Actual End',
    'Person Responsible'
]
for c, h in enumerate(schedule_headers, start=1):
    ws1.cell(row=11, column=c, value=h)
style_header(ws1, 11, 1, 10)

activities = [
    [1, 'Stakeholder Analysis & Mapping', 'Stakeholder Matrix', datetime(2026,3,1), datetime(2026,3,10), 25000, '-', datetime(2026,3,2), None, 'admin'],
    [2, 'Partnership Framework Design', 'Framework Document', datetime(2026,3,11), datetime(2026,4,5), 85000, 1, None, None, 'pm_user'],
    [3, 'Linkage Strategy Development', 'Strategy Report', datetime(2026,4,6), datetime(2026,5,15), 120000, 2, None, None, 'pm_user'],
    [4, 'Compliance & Internal Submission', 'Compliance Certificate', datetime(2026,5,16), datetime(2026,6,10), 45000, 3, None, None, 'recorder'],
    [5, 'Implementation of Shared Platform', 'Platform Go-Live', datetime(2026,6,11), datetime(2026,7,30), 250000, 4, None, None, 'pm_user'],
    [6, 'Final Review & Handover', 'Contract Completion', datetime(2026,8,1), datetime(2026,8,30), 50000, 5, None, None, 'admin']
]

for i, act in enumerate(activities):
    r = 12 + i
    for c, val in enumerate(act, start=1):
        cell = ws1.cell(r, c, val)
        cell.border = THIN_BORDER
        if isinstance(val, datetime):
            cell.number_format = 'YYYY-MM-DD'
        if c == 6:
            cell.number_format = 'R #,##0.00'

# --- 3. EXPENDITURE & RISK ---
def setup_sheet(name, headers, data):
    if name not in wb.sheetnames: wb.create_sheet(name)
    ws = wb[name]
    ws.delete_rows(1, 100)
    ws['A1'] = name.replace('_', ' ').upper()
    ws['A1'].font = Font(bold=True, size=14)
    for c, h in enumerate(headers, 1):
        ws.cell(4, c, h)
    style_header(ws, 4, 1, len(headers))
    for i, row in enumerate(data):
        r = 5 + i
        for c, val in enumerate(row, 1):
            cell = ws.cell(r, c, val)
            cell.border = THIN_BORDER
            if isinstance(val, datetime): cell.number_format = 'YYYY-MM-DD'
            if 'Amount' in headers[c-1]: cell.number_format = 'R #,##0.00'

setup_sheet('Expenditure_Log', ['Date', 'Activity ID', 'Category', 'Description', 'Reference (Invoice/PO)', 'Amount (R)'], [
    [datetime(2026,3,5), 1, 'Consultancy', 'Stakeholder meeting fees', 'INV-881', 12000.00],
    [datetime(2026,3,8), 1, 'Travel', 'Regional site visits', 'EXP-042', 4500.00],
])

setup_sheet('Risk_Register', ['Date Identified', 'Risk/Issue Description', 'Impact (H/M/L)', 'Status', 'Mitigation Action'], [
    [datetime(2026,3,1), 'Delay in stakeholder responses', 'M', 'Open', 'Send follow-up reminders'],
    [datetime(2026,3,4), 'Compliance framework update', 'H', 'Open', 'Consult with legal team'],
])

# Widths
ws1.column_dimensions['A'].width = 20
ws1.column_dimensions['B'].width = 35
ws1.column_dimensions['C'].width = 30
ws1.column_dimensions['E'].width = 20
ws1.column_dimensions['F'].width = 20

wb.save(output_file)
print(f"Created: {output_file}")
